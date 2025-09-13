"""XLSX helpers using openpyxl."""
from __future__ import annotations

from pathlib import Path
from typing import List

from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]

from console_io.csv_io import Task


def export_tasks_xlsx(path: str | Path, tasks: List[Task]) -> None:
    path = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.append(["id", "title", "owner"])
    for t in tasks:
        ws.append([t.id, t.title, t.owner])
    wb.save(path)


def import_tasks_xlsx(path: str | Path) -> List[Task]:
    path = Path(path)
    wb = load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    tasks = []
    for row in rows[1:]:
        data = dict(zip(header, row))
        tasks.append(Task(**data))
    return tasks
